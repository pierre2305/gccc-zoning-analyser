"""
GCCC Property Zoning Analyser
==============================
Scrapes property listings from realestate.com.au (Gold Coast),
geocodes each address, then queries the Gold Coast City Council
ArcGIS API to determine the zoning (LDR, MDR, etc.).

Requirements:
    pip install playwright requests pandas tqdm
    playwright install chromium

Usage:
    # Analyse listings scraped from realestate.com.au:
    python gccc_zoning_analyser.py --mode scrape --suburb "Burleigh Heads" --pages 3

    # Analyse a custom list of addresses from a text file (one per line):
    python gccc_zoning_analyser.py --mode file --input addresses.txt

    # Analyse a single address:
    python gccc_zoning_analyser.py --mode single --address "12 Smith Street, Burleigh Heads QLD 4220"

Output:
    zoning_results.csv  — one row per address with zone code, description, and density info
"""

import argparse
import csv
import json
import time
import sys
from datetime import datetime
from pathlib import Path

import requests
from tqdm import tqdm


# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

# GCCC ArcGIS — Residential Density overlay (V7, Layer 110)
# Server moved from maps.cityofgoldcoast.com.au → maps1.goldcoast.qld.gov.au
# This layer has the RESIDENTIAL_DENSITY field with LDR1/LDR2/RD1/RD2/RD3 codes
GCCC_DENSITY_URL = (
    "https://maps1.goldcoast.qld.gov.au/arcgis/rest/services"
    "/City_Plan_V7/MapServer/110/query"
)

# GCCC ArcGIS — V7 Zone layer (Layer 127)
# Use this if you want the raw zone category beyond residential density
GCCC_ZONE_URL = (
    "https://maps1.goldcoast.qld.gov.au/arcgis/rest/services"
    "/City_Plan_V7/MapServer/127/query"
)

# Nominatim geocoder (free fallback)
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_HEADERS = {"User-Agent": "GCCC-Zoning-Analyser/1.0 (personal-research)"}

# Queensland Government GeocodeServer (preferred — uses official council address data,
# returns property-centroid coordinates, no API key needed)
QLD_GEOCODE_URL = (
    "https://spatial-gis.information.qld.gov.au/arcgis/rest/services"
    "/Location/QldLocator/GeocodeServer/findAddressCandidates"
)

# realestate.com.au base search URL — sorted by newest listed first
# Format: /buy/in-{suburb}+qld+{postcode}/list-{page}?activeSort=list-date
REA_BASE_URL = "https://www.realestate.com.au/buy/in-{suburb_slug}/list-{page}?activeSort=list-date"

# Seconds to wait between requests (be polite to free services)
GEOCODE_DELAY = 1.2
ARCGIS_DELAY  = 0.3
SCRAPE_DELAY  = 5.0

# Human-readable zone descriptions
ZONE_DESCRIPTIONS = {
    "LDR1": "Low Density Residential — up to 12.5 dwellings/ha (1 per 800 m²)",
    "LDR2": "Low Density Residential — up to 16.6 dwellings/ha (1 per 600 m²)",
    "RD1":  "Medium Density — up to 25 dwellings/ha (1 per 400 m²)",
    "RD2":  "Medium Density — up to 33 dwellings/ha (1 per 300 m²)",
    "RD3":  "Higher Density — up to 40 dwellings/ha (1 per 250 m²)",
    "RD4":  "Higher Density — up to 50 dwellings/ha (1 per 200 m²)",
    "RD4A": "Higher Density — up to 66 dwellings/ha (1 per 150 m²)",
    "RD5":  "High Rise — up to 200 bedrooms/ha (1 bed per 50 m²)",
    "RD6":  "High Rise — up to 300 bedrooms/ha (1 bed per 33 m²)",
    "RD7":  "High Rise — up to 400 bedrooms/ha (1 bed per 25 m²)",
    "RD8":  "High Rise — up to 769 bedrooms/ha (1 bed per 13 m²)",
}

ZONE_CATEGORY = {
    "LDR1": "Low Density Residential",
    "LDR2": "Low Density Residential",
    "RD1":  "Medium Density Residential",
    "RD2":  "Medium Density Residential",
    "RD3":  "Higher Density Residential",
    "RD4":  "Higher Density Residential",
    "RD4A": "Higher Density Residential",
    "RD5":  "High Rise",
    "RD6":  "High Rise",
    "RD7":  "High Rise",
    "RD8":  "High Rise",
}


# ---------------------------------------------------------------------------
# STEP 1 — SCRAPE realestate.com.au
# ---------------------------------------------------------------------------

def scrape_listings(suburb: str, max_pages: int = 3) -> list[str]:
    """
    Use Playwright to scrape property addresses from realestate.com.au.
    Returns a list of address strings.

    NOTE: This respects realestate.com.au's robots.txt as best as possible by
    throttling requests. Review their ToS before running at scale.
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        print("ERROR: Playwright not installed. Run: pip install playwright && playwright install chromium")
        sys.exit(1)

    suburb_slug = suburb.lower()
    addresses = []

    print(f"\n🔍 Scraping realestate.com.au for '{suburb}' listings...")

    # Build the suburb slug — include postcode if provided e.g. "Burleigh Heads QLD 4220"
    # becomes "burleigh+heads,+qld+4220"
    suburb_slug = suburb_slug.replace(" ", "+").replace(",", "")
    # Ensure ,+qld is present for the search
    if "+qld" not in suburb_slug:
        suburb_slug = suburb_slug + ",+qld"

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-AU",
            timezone_id="Australia/Brisbane",
            viewport={"width": 1280, "height": 900},
            extra_http_headers={
                "Accept-Language": "en-AU,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            }
        )
        # Mask the webdriver property that bot detectors look for
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()

        for page_num in range(1, max_pages + 1):
            url = REA_BASE_URL.format(suburb_slug=suburb_slug, page=page_num)
            print(f"  Page {page_num}: {url}")

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
                # Wait for network to settle so JS-rendered listings appear
                try:
                    page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass  # networkidle timeout is fine — just means it's still loading ads etc
                time.sleep(SCRAPE_DELAY)  # wait for JS to render

                # Address structure confirmed from live REA HTML (May 2026):
                #
                #   <h2 class="residential-card__address-heading">
                #     <a class="details-link residential-card__details-link">
                #       <span class="">2 Tillys Place, Burleigh Heads</span>
                #     </a>
                #   </h2>
                #
                # Querying the <a> inside the h2 gives clean address text.
                # Fallback selectors cover older/alternate card layouts.
                selectors = [
                    "h2.residential-card__address-heading a",   # confirmed live May 2026
                    "h2.residential-card__address-heading span",
                    '[data-testid="listing-card-address"]',
                    '[data-testid="address-line1"]',
                    'span[itemprop="streetAddress"]',
                    '.property-info-address',
                ]

                found = False
                for selector in selectors:
                    try:
                        page.wait_for_selector(selector, timeout=5000)
                        cards = page.query_selector_all(selector)
                        if cards:
                            page_addresses = []
                            for card in cards:
                                addr = card.inner_text().strip()
                                # A real address contains at least one digit
                                if addr and any(c.isdigit() for c in addr) and addr not in addresses:
                                    addresses.append(addr)
                                    page_addresses.append(addr)
                            if page_addresses:
                                print(f"    ✓ Found {len(page_addresses)} listings (selector: {selector})")
                                found = True
                                break
                    except PWTimeout:
                        continue

                if not found:
                    title = page.title()
                    print(f"    ⚠️  No listings found on page {page_num} (page title: '{title}')")
                    print(f"    URL tried: {url}")

                    # Save raw HTML so we can inspect what the browser actually got
                    debug_file = f"rea_debug_page{page_num}.html"
                    with open(debug_file, "w", encoding="utf-8") as dbf:
                        dbf.write(page.content())
                    print(f"    📄 Raw HTML saved to: {debug_file}")
                    print( "    Open that file in a text editor and search for a known")
                    print( "    address from the page to find the right CSS selector.")

                # Stop early if last page returned nothing
                if not found and page_num > 1:
                    print(f"    Stopping — no more listings found after page {page_num - 1}.")
                    break

            except Exception as e:
                print(f"    ⚠️  Error on page {page_num}: {e}")

            time.sleep(SCRAPE_DELAY)

        browser.close()

    print(f"\n  ✅ Scraped {len(addresses)} unique addresses\n")
    return addresses



def scrape_domain(suburb: str, max_pages: int = 3) -> list[str]:
    """
    Scrape property listings from Domain.com.au (less bot-protected than REA).
    Domain uses server-side rendering so addresses are in the raw HTML.
    Falls back gracefully if structure changes.
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        print("ERROR: Playwright not installed.")
        sys.exit(1)

    # Domain URL format: /sale/suburb-state-postcode/
    # e.g. /sale/burleigh-heads-qld-4220/?sort=dateupdated
    slug = suburb.lower().replace(" ", "-")
    # Add qld if not present
    if "-qld" not in slug:
        slug = slug + "-qld"

    base = f"https://www.domain.com.au/sale/{slug}/?sort=dateupdated&page="
    addresses = []

    print(f"\n🔍 Scraping Domain.com.au for '{suburb}' listings...")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-AU",
            timezone_id="Australia/Brisbane",
            viewport={"width": 1280, "height": 900},
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()

        for page_num in range(1, max_pages + 1):
            url = base + str(page_num)
            print(f"  Page {page_num}: {url}")
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
                try:
                    page.wait_for_load_state("networkidle", timeout=8000)
                except Exception:
                    pass
                time.sleep(SCRAPE_DELAY)

                # Domain address selectors (confirmed structure)
                selectors = [
                    '[data-testid="listing-card-address"]',
                    '[data-testid="address"]',
                    "h2.listing-card__content-copy--address",
                    '[class*="address__Address"]',
                    "span.css-bqbbuf",       # Domain styled-component address span
                    "address",
                ]

                found = False
                for selector in selectors:
                    try:
                        page.wait_for_selector(selector, timeout=5000)
                        cards = page.query_selector_all(selector)
                        if cards:
                            page_addrs = []
                            for card in cards:
                                addr = card.inner_text().strip()
                                if addr and any(c.isdigit() for c in addr) and addr not in addresses:
                                    addresses.append(addr)
                                    page_addrs.append(addr)
                            if page_addrs:
                                print(f"    ✓ Found {len(page_addrs)} listings (selector: {selector})")
                                found = True
                                break
                    except PWTimeout:
                        continue

                if not found:
                    title = page.title()
                    print(f"    ⚠️  No listings found on page {page_num} (title: '{title}')")
                    debug_file = f"domain_debug_page{page_num}.html"
                    with open(debug_file, "w", encoding="utf-8") as dbf:
                        dbf.write(page.content())
                    print(f"    📄 Debug HTML saved to: {debug_file}")
                    if page_num > 1:
                        break

            except Exception as e:
                print(f"    ⚠️  Error on page {page_num}: {e}")
            time.sleep(SCRAPE_DELAY)

        browser.close()

    print(f"\n  ✅ Scraped {len(addresses)} unique addresses from Domain\n")
    return addresses

def load_addresses_from_file(filepath: str) -> list[dict]:
    """
    Load addresses from file. Supports:
    - CSV with 'address' column (and optional 'price' column from harvester)
    - Plain .txt file with one address per line
    Returns a list of dicts with at least an 'address' key.
    """
    import csv as _csv
    path = Path(filepath)
    if not path.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    records = []

    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            for row in reader:
                addr = row.get("address", "").strip()
                if addr:
                    records.append({
                        "address": addr,
                        "price":   row.get("price", "").strip(),
                        "url":     row.get("url", "").strip(),
                    })
    else:
        # Plain text — one address per line
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                records.append({"address": line, "price": "", "url": ""})

    print(f"📂 Loaded {len(records)} listings from {filepath}")
    return records


# ---------------------------------------------------------------------------
# STEP 2 — GEOCODE
# ---------------------------------------------------------------------------

def geocode_address(address: str, state_hint: str = "QLD", country: str = "Australia") -> tuple[float | None, float | None]:
    """
    Geocode an address to (lat, lng).

    Priority order:
    1. Google Maps API (if GOOGLE_MAPS_API_KEY env var is set)
    2. Queensland Government GeocodeServer (free, uses official council data,
       returns property-centroid coordinates — most accurate for QLD addresses)
    3. Nominatim / OpenStreetMap (free fallback, less accurate)
    """
    import os

    google_key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if google_key:
        return _geocode_google(address, google_key)

    # Try QLD Government geocoder first (best for Gold Coast addresses)
    result = _geocode_qld(address)
    if result != (None, None):
        return result

    # Fall back to Nominatim
    return _geocode_nominatim(address, state_hint, country)


def _geocode_qld(address: str) -> tuple[float | None, float | None]:
    """
    Geocode using the Queensland Government ArcGIS GeocodeServer.
    Free, no API key. Uses QAMF (Queensland Address Management Framework)
    data supplied by local councils — returns property-centroid coordinates.
    """
    params = {
        "SingleLine": address,
        "outFields":  "Match_addr,Score",
        "outSR":      "4326",      # return WGS84 lat/lng
        "maxLocations": 1,
        "f":          "json",
    }
    try:
        resp = requests.get(QLD_GEOCODE_URL, params=params, timeout=10)
        resp.raise_for_status()
        candidates = resp.json().get("candidates", [])
        if candidates and candidates[0].get("score", 0) >= 80:
            loc = candidates[0]["location"]
            return float(loc["y"]), float(loc["x"])  # y=lat, x=lng
    except Exception as e:
        print(f"    ⚠️  QLD geocoder error for '{address}': {e}")
    return None, None


def _geocode_nominatim(address: str, state_hint: str = "QLD", country: str = "Australia") -> tuple[float | None, float | None]:
    """Geocode using OpenStreetMap Nominatim (free fallback)."""
    query = f"{address}, {state_hint}, {country}"
    params = {"q": query, "format": "json", "limit": 1, "countrycodes": "au"}
    try:
        resp = requests.get(NOMINATIM_URL, params=params, headers=NOMINATIM_HEADERS, timeout=10)
        resp.raise_for_status()
        results = resp.json()
        if results:
            return float(results[0]["lat"]), float(results[0]["lon"])
    except Exception as e:
        print(f"    ⚠️  Nominatim geocoding error for '{address}': {e}")
    return None, None


def _geocode_google(address: str, api_key: str) -> tuple[float | None, float | None]:
    """Geocode using Google Maps API (more accurate for Australian addresses)."""
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": address + ", Australia", "key": api_key}

    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        if data.get("status") == "OK" and data.get("results"):
            loc = data["results"][0]["geometry"]["location"]
            return loc["lat"], loc["lng"]
    except Exception as e:
        print(f"    ⚠️  Google geocoding error for '{address}': {e}")

    return None, None


# ---------------------------------------------------------------------------
# STEP 3 — QUERY GCCC ARCGIS API
# ---------------------------------------------------------------------------

def get_residential_density(lat: float, lng: float) -> dict:
    """
    Two-step zone lookup for a WGS84 coordinate:

    Step 1 — Zone layer (Layer 127): covers EVERY property with plain-English
    zone names like "Low density residential", "Medium density residential".
    This is the primary and reliable source.

    Step 2 — Residential Density overlay (Layer 110): only exists for streets
    where a specific density code (LDR1, LDR2, RD1-RD8) has been mapped.
    Gives a more precise code where available, otherwise left blank.

    inSR=4326 tells ArcGIS our input is WGS84 — no coordinate conversion needed.
    """
    common_params = {
        "geometry":       f"{lng},{lat}",   # ArcGIS uses x,y = lng,lat order
        "geometryType":   "esriGeometryPoint",
        "inSR":           "4326",           # input is WGS84 — server reprojects
        "spatialRel":     "esriSpatialRelIntersects",
        "returnGeometry": "false",
        "f":              "json",
    }

    # ------------------------------------------------------------------
    # Step 1 — Zone layer (primary, covers all properties)
    # Fields: LVL1_ZONE = zone name, ZONE = zone + precinct combined
    # ------------------------------------------------------------------
    zone_name     = ""
    zone_precinct = ""
    zone_params = {**common_params,
                   "outFields": "ZONE,LVL1_ZONE,ZONE_PRECINCT",
                   "distance":  "20",    # 20m buffer — catches points on road/boundary
                   "units":     "esriSRUnit_Meter"}
    # Retry up to 3 times with increasing timeout
    for attempt, timeout_secs in enumerate([15, 25, 40], start=1):
        try:
            resp = requests.get(GCCC_ZONE_URL, params=zone_params, timeout=timeout_secs)
            resp.raise_for_status()
            features = resp.json().get("features", [])
            if features:
                attrs         = features[0]["attributes"]
                zone_name     = attrs.get("LVL1_ZONE") or attrs.get("ZONE") or ""
                zone_precinct = attrs.get("ZONE_PRECINCT") or ""
            break  # success — exit retry loop
        except requests.exceptions.Timeout:
            if attempt == 3:
                return _empty_zone_result("API timeout — GCCC server not responding after 3 attempts")
            print(f"    ⏱  Zone API timeout (attempt {attempt}/3) — retrying in 2s...")
            time.sleep(2)  # brief pause before retry
            continue
        except Exception as e:
            return _empty_zone_result(str(e))

    if not zone_name:
        return _empty_zone_result("Address is outside GCCC boundary or unzoned")

    # ------------------------------------------------------------------
    # Step 2 — Residential Density overlay (optional, not all properties)
    # Fields: RESIDENTIAL_DENSITY (e.g. "LDR1"), CAT_DESC, OVL_CAT
    # ------------------------------------------------------------------
    density_code = ""
    cat_desc     = ""
    ovl_cat      = ""
    ovl2_desc    = ""
    try:
        params = {**common_params, "outFields": "RESIDENTIAL_DENSITY,CAT_DESC,OVL_CAT,OVL2_DESC"}
        resp = requests.get(GCCC_DENSITY_URL, params=params, timeout=30)
        resp.raise_for_status()
        features = resp.json().get("features", [])
        if features:
            attrs        = features[0]["attributes"]
            density_code = attrs.get("RESIDENTIAL_DENSITY") or ""
            cat_desc     = attrs.get("CAT_DESC") or ""
            ovl_cat      = attrs.get("OVL_CAT") or ""
            ovl2_desc    = attrs.get("OVL2_DESC") or ""
    except Exception:
        pass  # density overlay is optional — don't fail if it errors

    # ------------------------------------------------------------------
    # Combine both sources into a single result
    # ------------------------------------------------------------------
    zone_cat = zone_name.title()  # "Low density residential" -> "Low Density Residential"

    if density_code and density_code in ZONE_DESCRIPTIONS:
        zone_desc = ZONE_DESCRIPTIONS[density_code]
    else:
        zone_desc = zone_name

    return {
        "zone_name":     zone_name,
        "zone_cat":      zone_cat,
        "zone_precinct": zone_precinct,
        "zone_desc":     zone_desc,
        "zone_code":     density_code,   # blank if not in density overlay
        "cat_desc":      cat_desc,
        "ovl_cat":       ovl_cat,
        "ovl2_desc":     ovl2_desc,
        "api_error":     None,
    }


def _empty_zone_result(reason: str) -> dict:
    return {
        "zone_name":     "",
        "zone_cat":      "Lookup failed",
        "zone_precinct": "",
        "zone_desc":     reason,
        "zone_code":     "",
        "cat_desc":      "",
        "ovl_cat":       "",
        "ovl2_desc":     "",
        "api_error":     reason,
    }


# ---------------------------------------------------------------------------
# STEP 4 — ORCHESTRATE & OUTPUT
# ---------------------------------------------------------------------------

def analyse_addresses(addresses: list[dict], output_file: str = "zoning_results.csv") -> list[dict]:
    """
    Main pipeline: geocode each address then query GCCC for its zoning.
    Writes results to CSV and returns the list of result dicts.
    """
    results = []
    failed_geocode = []
    failed_zoning  = []

    print(f"\n{'='*60}")
    print(f"  Analysing {len(addresses)} addresses")
    print(f"  Output: {output_file}")
    print(f"{'='*60}\n")

    for item in tqdm(addresses, desc="Processing", unit="addr"):
        address = item["address"] if isinstance(item, dict) else item
        price   = item.get("price", "") if isinstance(item, dict) else ""
        url     = item.get("url", "") if isinstance(item, dict) else ""
        result  = {"address": address, "price": price, "url": url}

        # --- Geocode ---
        time.sleep(GEOCODE_DELAY)
        lat, lng = geocode_address(address)

        if lat is None or lng is None:
            tqdm.write(f"  ✗ Geocode failed: {address}")
            failed_geocode.append(address)
            result.update({
                "lat": "", "lng": "",
                **_empty_zone_result("Geocoding failed"),
            })
            results.append(result)
            continue

        result["lat"] = round(lat, 6)
        result["lng"] = round(lng, 6)

        # --- Zone lookup ---
        time.sleep(ARCGIS_DELAY)
        zone = get_residential_density(lat, lng)
        result.update(zone)

        if zone["api_error"]:
            tqdm.write(f"  ✗ Zone lookup failed: {address} — {zone['api_error']}")
            failed_zoning.append(address)
        else:
            tqdm.write(
                f"  ✓ {address[:50]:<50} → {zone['zone_code'] or '?':>5}  {zone['zone_cat']}"
            )

        results.append(result)

    # --- Write CSV ---
    _write_csv(results, output_file)
    write_excel(results, output_file)

    # --- Summary ---
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  Total processed:    {len(results)}")
    print(f"  Geocode failures:   {len(failed_geocode)}")
    print(f"  Zone lookup failed: {len(failed_zoning)}")
    print()

    zone_counts = {}
    for r in results:
        cat = r.get("zone_cat", "Unknown")
        zone_counts[cat] = zone_counts.get(cat, 0) + 1

    print("  Zone breakdown:")
    for cat, count in sorted(zone_counts.items()):
        print(f"    {cat:<35} {count:>4} properties")

    print(f"\n  Results saved to: {output_file}")

    return results



# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

# Row fill colours by zone category
ZONE_COLOURS = {
    "Low Density Residential":    "C6EFCE",   # light green
    "Medium Density Residential": "FFECC7",   # light orange
    "Higher Density Residential": "FFC7CE",   # light red/pink
    "High Rise":                  "E2CFFE",   # light purple
    "Lookup failed":              "D9D9D9",   # light grey
}
ZONE_COLOUR_DEFAULT = "FFFFFF"               # white for anything else

# Friendly column headers for the Excel sheet
EXCEL_HEADERS = {
    "address":        "Address",
    "price":          "Price",
    "url":            "View Listing",
    "zone_name":      "Zone Name",
    "zone_cat":       "Zone Category",
    "zone_precinct":  "Zone Precinct",
    "zone_code":      "Density Code",
    "zone_desc":      "Density Description",
    "cat_desc":       "GCCC Category",
    "ovl_cat":        "Overlay Cat",
    "ovl2_desc":      "Overlay 2 Desc",
    "api_error":      "Error",
    "lat":            "Latitude",
    "lng":            "Longitude",
}


def write_excel(results: list[dict], filepath: str) -> None:
    """
    Write results to a colour-coded, sorted Excel file.
    - Sorted by zone_cat (Low Density → Medium → Higher → High Rise → Other)
    - Each row filled with the zone category colour
    - Price column forced to text / left-aligned
    - Header row bold, frozen, auto-filtered
    - Column widths auto-fitted
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️  openpyxl not installed — skipping Excel output.")
        print("       Run: pip install openpyxl")
        return

    if not results:
        return

    # Sort order for zone categories
    sort_order = {
        "Low Density Residential":    0,
        "Medium Density Residential": 1,
        "Higher Density Residential": 2,
        "High Rise":                  3,
        "Lookup failed":              5,
    }
    sorted_results = sorted(
        results,
        key=lambda r: (sort_order.get(r.get("zone_cat", ""), 4), r.get("address", ""))
    )

    fieldnames = [
        "address", "price", "url",
        "zone_name", "zone_cat", "zone_precinct",
        "zone_code", "zone_desc",
        "cat_desc", "ovl_cat", "ovl2_desc",
        "api_error",
        "lat", "lng",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zoning Results"

    # --- Styles ---
    header_font    = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill    = PatternFill("solid", start_color="2F4F6F")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font      = Font(name="Arial", size=10)
    left_align     = Alignment(horizontal="left",   vertical="center")
    center_align   = Alignment(horizontal="center", vertical="center")
    thin_side      = Side(style="thin", color="D0D0D0")
    thin_border    = Border(left=thin_side, right=thin_side,
                            top=thin_side,  bottom=thin_side)

    # --- Header row ---
    headers = [EXCEL_HEADERS.get(f, f) for f in fieldnames]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[1].height = 30

    # --- Data rows ---
    price_col_idx = fieldnames.index("price") + 1  # 1-based
    url_col_idx   = fieldnames.index("url")   + 1  # 1-based

    for row_data in sorted_results:
        row_values = []
        for field in fieldnames:
            val = row_data.get(field, "") or ""
            # Force price to plain string so Excel never auto-converts to currency
            if field == "price":
                val = str(val)
            row_values.append(val)
        ws.append(row_values)

        row_idx = ws.max_row
        zone_cat = row_data.get("zone_cat", "")
        fill_hex  = ZONE_COLOURS.get(zone_cat, ZONE_COLOUR_DEFAULT)
        row_fill  = PatternFill("solid", start_color=fill_hex)

        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font   = data_font
            cell.fill   = row_fill
            cell.border = thin_border
            cell.alignment = left_align
            # Price column: force text format so $ amounts stay left-aligned strings
            if col_idx == price_col_idx:
                cell.number_format = "@"
            # URL column: render as a clickable hyperlink labelled "View"
            if col_idx == url_col_idx and cell.value:
                url_val = cell.value
                cell.value     = "View listing"
                cell.hyperlink = url_val
                cell.font      = Font(name="Arial", size=10,
                                      color="0563C1", underline="single")

    # --- Auto-fit column widths ---
    for col_idx, field in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)
        header_len = len(EXCEL_HEADERS.get(field, field))
        max_len    = header_len
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        # Cap width and add a little padding
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # --- Freeze header row + auto-filter ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Colour legend sheet ---
    legend = wb.create_sheet("Legend")
    legend["A1"] = "Zone Category"
    legend["B1"] = "Row Colour"
    legend["A1"].font = Font(name="Arial", bold=True, size=10)
    legend["B1"].font = Font(name="Arial", bold=True, size=10)
    legend_data = [
        ("Low Density Residential",    ZONE_COLOURS["Low Density Residential"]),
        ("Medium Density Residential", ZONE_COLOURS["Medium Density Residential"]),
        ("Higher Density Residential", ZONE_COLOURS["Higher Density Residential"]),
        ("High Rise",                  ZONE_COLOURS["High Rise"]),
        ("Lookup failed / Other",      ZONE_COLOURS["Lookup failed"]),
    ]
    for i, (cat, colour) in enumerate(legend_data, start=2):
        legend.cell(row=i, column=1, value=cat).font = Font(name="Arial", size=10)
        colour_cell = legend.cell(row=i, column=2, value="")
        colour_cell.fill = PatternFill("solid", start_color=colour)
    legend.column_dimensions["A"].width = 32
    legend.column_dimensions["B"].width = 14

    xlsx_path = filepath.replace(".csv", ".xlsx")
    if not xlsx_path.endswith(".xlsx"):
        xlsx_path += ".xlsx"
    wb.save(xlsx_path)
    print(f"  📊 Excel file saved to: {xlsx_path}")



def write_excel_bytes(results: list[dict]) -> bytes | None:
    """
    Same as write_excel() but returns the file as bytes instead of
    writing to disk. Used by the Streamlit UI for download buttons.
    """
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    # Reuse all the same logic — write to a BytesIO buffer
    import tempfile, os
    tmp = tempfile.mktemp(suffix=".xlsx")
    write_excel(results, tmp.replace(".xlsx", ".csv"))  # write_excel appends .xlsx
    xlsx_path = tmp.replace(".xlsx", "") + ".xlsx"
    if not os.path.exists(xlsx_path):
        xlsx_path = tmp
    try:
        with open(xlsx_path, "rb") as f:
            return f.read()
    except FileNotFoundError:
        return None

def _write_csv(results: list[dict], filepath: str) -> None:
    if not results:
        print("  ⚠️  No results to write.")
        return

    fieldnames = [
        "address", "price", "url",
        "zone_name", "zone_cat", "zone_precinct",
        "zone_code", "zone_desc",
        "cat_desc", "ovl_cat", "ovl2_desc",
        "api_error",
        "lat", "lng",
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)


# ---------------------------------------------------------------------------
# STEP 5 — OPTIONAL: FILTER & REPORT
# ---------------------------------------------------------------------------

def filter_by_zone(results: list[dict], zone_cat: str) -> list[dict]:
    """
    Filter results to a specific zone category.
    zone_cat examples: 'Low Density Residential', 'Medium Density Residential'
    """
    return [r for r in results if r.get("zone_cat") == zone_cat]


def print_summary_table(results: list[dict]) -> None:
    """Pretty-print a summary table to stdout."""
    col_w = [52, 6, 30]
    header = f"{'Address':<{col_w[0]}} {'Zone':<{col_w[1]}} {'Category':<{col_w[2]}}"
    print("\n" + header)
    print("-" * sum(col_w))
    for r in results:
        addr = r["address"][:col_w[0]]
        code = r.get("zone_code", "?")[:col_w[1]]
        cat  = r.get("zone_cat", "?")[:col_w[2]]
        print(f"{addr:<{col_w[0]}} {code:<{col_w[1]}} {cat:<{col_w[2]}}")


# ---------------------------------------------------------------------------
# CLI ENTRY POINT
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Analyse Gold Coast property listings against GCCC zoning data.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scrape realestate.com.au for a suburb:
  python gccc_zoning_analyser.py --mode scrape --suburb "Burleigh Heads" --pages 3

  # Read addresses from a text file:
  python gccc_zoning_analyser.py --mode file --input my_addresses.txt

  # Single address:
  python gccc_zoning_analyser.py --mode single --address "12 Smith St, Burleigh Heads QLD 4220"

  # Use Google Maps geocoder (more accurate):
  export GOOGLE_MAPS_API_KEY=your_key_here
  python gccc_zoning_analyser.py --mode file --input my_addresses.txt

Tip: For the scrape mode, you need Playwright:
  pip install playwright && playwright install chromium
        """,
    )

    parser.add_argument(
        "--mode", choices=["scrape", "file", "single"], required=True,
        help="Source of addresses to analyse"
    )
    parser.add_argument(
        "--source", choices=["domain", "rea"], default="domain",
        help="Which site to scrape in scrape mode (default: domain)"
    )
    parser.add_argument(
        "--suburb", default="Burleigh Heads",
        help="Suburb name for scrape mode (default: Burleigh Heads)"
    )
    parser.add_argument(
        "--pages", type=int, default=3,
        help="Number of pages to scrape from realestate.com.au (default: 3)"
    )
    parser.add_argument(
        "--input", default="addresses.txt",
        help="Path to text file with addresses, one per line (file mode)"
    )
    parser.add_argument(
        "--address", default="",
        help="Single address to analyse (single mode)"
    )
    parser.add_argument(
        "--output", default="zoning_results.csv",
        help="Output CSV file path (default: zoning_results.csv)"
    )
    parser.add_argument(
        "--filter-ldr", action="store_true",
        help="After analysis, print only Low Density Residential results"
    )
    parser.add_argument(
        "--filter-mdr", action="store_true",
        help="After analysis, print only Medium Density Residential results"
    )

    args = parser.parse_args()

    # --- Gather addresses ---
    if args.mode == "scrape":
        if args.source == "domain":
            addresses = scrape_domain(args.suburb, args.pages)
        else:
            addresses = scrape_listings(args.suburb, args.pages)
    elif args.mode == "file":
        addresses = load_addresses_from_file(args.input)
    elif args.mode == "single":
        if not args.address:
            print("ERROR: --address is required for single mode")
            sys.exit(1)
        addresses = [args.address]

    if not addresses:
        print("No addresses to process. Exiting.")
        sys.exit(0)

    # --- Run analysis ---
    results = analyse_addresses(addresses, args.output)

    # --- Optional filters ---
    if args.filter_ldr:
        ldr = filter_by_zone(results, "Low Density Residential")
        print(f"\n  Low Density Residential ({len(ldr)} properties):")
        print_summary_table(ldr)

    if args.filter_mdr:
        mdr = filter_by_zone(results, "Medium Density Residential")
        print(f"\n  Medium Density Residential ({len(mdr)} properties):")
        print_summary_table(mdr)


if __name__ == "__main__":
    main()
